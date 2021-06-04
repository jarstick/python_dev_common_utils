# -*- coding: utf-8 -*-

import datetime
import os
from shutil import copy

import openpyxl
import pandas as pd
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.colors import BLACK, RED
from openpyxl.styles.fonts import Font

from utils.log import log


class WriteExcel:
    """写入Excel"""

    now_time = datetime.datetime.now().strftime("%Y年%m月%d日%H时%M分")

    def __init__(self, old_file, sheet_name='Sheet1'):
        """如果在初始化的时候不传入文件路径，则使用项目的测试用例默认路径并copy"""
        self.result_file_name = copy(old_file, os.path.dirname(os.path.dirname(
            os.path.dirname(__file__))) + os.sep + "excel_result" + os.sep + self.now_time + '_copy.xlsx',
                                     follow_symlinks=True)
        self.result_book = openpyxl.open(self.result_file_name)
        self.sheet_name = sheet_name
        self.sheet = self.result_book[sheet_name]

    def write_result(self, row=0, col=0, value=''):
        align = Alignment(horizontal='center', vertical='center')
        self.sheet.cell(row=row, column=col).alignment = align
        if value == "Pass" or value == "pass":
            # 0 黑色
            success_font = Font(color=BLACK)
            self.sheet.cell(row, col, "Pass").font = success_font

            log.info('在%d行%d列成功写入测试结果:%s' % (row, col, value))
        elif value == "fail" or value == "Fail":
            # 2 红色
            fail_font = Font(name='Arial', size=10, color=RED)
            self.sheet.cell(row, col, "Fail").font = fail_font
            log.info('在%d行%d列成功写入测试结果:%s' % (row, col, value))
        else:
            self.sheet.cell(row, col, value)
            log.info('在%d行%d列成功写入测试结果:%s' % (row, col, value))
        self.result_book.save(self.result_file_name)

    def get_row_num(self, case_id):
        """
        根据对应的case_id找到对应的行号
        """
        num = 0
        clols_data = self.get_cols_data()
        for col_data in clols_data:
            if case_id in col_data:
                return num
            num += 1

    def get_cols_data(self, col_id=None):
        """
        获取某一列的内容
        """
        if col_id is not None:
            cols = self.sheet.col_values(col_id)
        else:
            cols = self.sheet.col_values(0)
        return cols

    def get_current_rownum(self, col_name, cell_value):
        df = pd.read_excel(self.result_file_name)
        return int(df[df[col_name].isin([cell_value])].index[0] + 2)
